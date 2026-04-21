#!/usr/bin/env python3
"""Lead generation scraper: Google Maps -> Hunter.io -> website emails -> Excel."""

import argparse
import os
import re
import sys
import time
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
HUNTER_ENDPOINT = "https://api.hunter.io/v2/domain-search"

JUNK_EMAIL_PREFIXES = {
    "noreply", "no-reply", "donotreply", "do-not-reply",
    "support", "help", "info", "hello", "contact", "sales",
    "billing", "admin", "webmaster", "hostmaster", "postmaster",
    "abuse", "privacy", "legal", "careers", "jobs", "hr",
    "marketing", "newsletter", "notifications", "alerts", "team",
    "office", "media", "press", "service", "services",
}

RETRY_BACKOFF = [1, 3, 7]

FIELDS = [
    "Business Name",
    "Phone",
    "Email",
    "Website",
    "City",
    "Review Count",
    "Qualified",
    "Notes",
]


def scrape_google_maps(niche: str, city: str, max_results: int = 40, headless: bool = True):
    """Search Google Maps and extract business name, phone, website, address, review count."""
    query = f"{niche} in {city}"
    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(locale="en-US")
        page = ctx.new_page()
        page.goto(f"https://www.google.com/maps/search/{query.replace(' ', '+')}", timeout=60000)

        _dismiss_consent(page)

        try:
            page.wait_for_selector('div[role="feed"]', timeout=20000)
        except Exception:
            print("Could not load results feed.", file=sys.stderr)
            browser.close()
            return results

        feed = page.locator('div[role="feed"]')
        prev_count = 0
        stagnant = 0
        while stagnant < 3:
            feed.evaluate("el => el.scrollBy(0, el.scrollHeight)")
            time.sleep(1.8)
            cards = page.locator('div[role="feed"] > div > div[jsaction]')
            count = cards.count()
            if count == prev_count:
                stagnant += 1
            else:
                stagnant = 0
            prev_count = count
            if count >= max_results:
                break

        cards = page.locator('div[role="feed"] a.hfpxzc')
        total = min(cards.count(), max_results)
        print(f"Found {total} cards. Extracting details...")

        for i in range(total):
            try:
                card = cards.nth(i)
                card.scroll_into_view_if_needed()
                card.click()
                page.wait_for_selector('h1.DUwDvf', timeout=10000)
                time.sleep(1.0)

                name = _safe_text(page, 'h1.DUwDvf')
                review_count = _extract_review_count(page)
                address = _extract_info(page, 'button[data-item-id="address"]')
                phone = _extract_info(page, 'button[data-item-id^="phone"]')
                website = _extract_info(page, 'a[data-item-id="authority"]', attr="href")

                results.append({
                    "name": name,
                    "phone": phone,
                    "website": website,
                    "address": address,
                    "review_count": review_count,
                })
            except Exception as e:
                print(f"  [{i}] skipped: {e}", file=sys.stderr)
                continue

        browser.close()
    return results


def _dismiss_consent(page):
    """Click through Google's consent/cookie wall if it appears."""
    consent_selectors = [
        'button[aria-label="Accept all"]',
        'button[aria-label="Agree to the use of cookies and other data for the purposes described"]',
        'form[action*="consent"] button[type="submit"]',
        '#L2AGLb',  # legacy Google consent button id
    ]
    for sel in consent_selectors:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=2000):
                btn.click()
                time.sleep(1.0)
                return
        except Exception:
            continue


def _get_with_retry(url: str, headers: dict = None, params: dict = None, timeout: int = 15):
    """GET with exponential backoff. Returns requests.Response or raises on final failure."""
    last_exc = None
    for attempt, wait in enumerate(RETRY_BACKOFF):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=timeout)
            return r
        except Exception as e:
            last_exc = e
            if attempt < len(RETRY_BACKOFF) - 1:
                time.sleep(wait)
    raise last_exc


def _safe_text(page, selector):
    try:
        return page.locator(selector).first.inner_text().strip()
    except Exception:
        return ""


def _extract_info(page, selector, attr=None):
    try:
        loc = page.locator(selector).first
        if attr:
            return loc.get_attribute(attr) or ""
        text = loc.get_attribute("aria-label") or loc.inner_text()
        return re.sub(r"^(Phone:|Address:|Website:)\s*", "", text).strip()
    except Exception:
        return ""


def _extract_review_count(page):
    try:
        label = page.locator('button[jsaction*="pane.rating.moreReviews"]').first.inner_text()
        digits = re.sub(r"[^\d]", "", label)
        return int(digits) if digits else 0
    except Exception:
        try:
            label = page.locator('span[aria-label*="review"]').first.get_attribute("aria-label") or ""
            digits = re.sub(r"[^\d]", "", label.split("review")[0])
            return int(digits) if digits else 0
        except Exception:
            return 0


def hunter_domain_search(domain: str, api_key: str):
    """Return (email, notes) for the most likely owner/contact at a domain."""
    if not domain or not api_key:
        return "", ""
    try:
        r = _get_with_retry(
            HUNTER_ENDPOINT,
            params={"domain": domain, "api_key": api_key, "limit": 10},
            timeout=20,
        )
        if r.status_code != 200:
            return "", f"hunter:{r.status_code}"
        data = r.json().get("data", {})
        emails = data.get("emails", []) or []
        if not emails:
            return "", ""

        priority = {"owner": 0, "ceo": 0, "founder": 0, "president": 1, "manager": 2}
        def score(e):
            pos = (e.get("position") or "").lower()
            dept = (e.get("department") or "").lower()
            for key, val in priority.items():
                if key in pos:
                    return val
            if dept in ("executive", "management"):
                return 3
            return 5
        emails.sort(key=score)
        best = emails[0]
        email = best.get("value", "")
        note = (best.get("position") or "").strip()
        return email, f"hunter:{note}" if note else "hunter"
    except Exception as e:
        return "", f"hunter_err:{type(e).__name__}"


def scrape_website_emails(url: str):
    """Fetch home + common contact pages, return first email found."""
    if not url:
        return "", ""
    try:
        parsed = urlparse(url)
        base = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme else f"https://{url}"
    except Exception:
        return "", ""

    candidates = [url, f"{base}/contact", f"{base}/contact-us", f"{base}/about"]
    seen = set()
    headers = {"User-Agent": "Mozilla/5.0 (compatible; LeadBot/1.0)"}
    for link in candidates:
        if link in seen:
            continue
        seen.add(link)
        try:
            r = _get_with_retry(link, headers=headers, timeout=15)
            if r.status_code != 200:
                continue
            found = EMAIL_RE.findall(r.text)
            found = [
                e for e in found
                if not e.lower().endswith((".png", ".jpg", ".gif", ".svg", ".webp"))
                and e.split("@")[0].lower() not in JUNK_EMAIL_PREFIXES
            ]
            if found:
                return found[0], f"site:{urlparse(link).path or '/'}"
        except Exception:
            continue
    return "", ""


def domain_from_url(url: str) -> str:
    if not url:
        return ""
    try:
        netloc = urlparse(url if "://" in url else f"https://{url}").netloc
        return netloc.lower().removeprefix("www.")
    except Exception:
        return ""


def qualify(review_count: int, extra_notes: str) -> tuple[str, str]:
    notes = []
    high_volume = review_count >= 50
    if high_volume:
        notes.append(f"high_volume({review_count}_reviews)")
    if extra_notes:
        notes.append(extra_notes)
    qualified = "yes" if high_volume else "no"
    return qualified, "; ".join(n for n in notes if n)


def _dedup_key(name: str, phone: str) -> tuple:
    return (name.strip().lower(), re.sub(r"\D", "", phone or ""))


def load_existing_keys(xlsx_path: Path) -> set:
    """Return the set of (name_lower, phone_digits) tuples already in the workbook."""
    keys = set()
    if not xlsx_path.exists():
        return keys
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        name_i = headers.index("Business Name")
        phone_i = headers.index("Phone")
    except ValueError:
        return keys
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[name_i]:
            keys.add(_dedup_key(str(row[name_i]), str(row[phone_i] or "")))
    return keys


def _new_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(FIELDS)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    widths = [32, 18, 32, 36, 18, 14, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    return wb


def write_rows(xlsx_path: Path, rows: list[dict]):
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = _new_workbook(xlsx_path)
        ws = wb.active
    qualified_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    for row in rows:
        ws.append([row[f] for f in FIELDS])
        if row["Qualified"] == "yes":
            for cell in ws[ws.max_row]:
                cell.fill = qualified_fill
    wb.save(xlsx_path)


def run(niche: str, city: str, output: Path, hunter_key: str, max_results: int, headless: bool):
    existing = load_existing_keys(output)
    print(f"Master workbook: {output} ({len(existing)} existing rows)")

    businesses = scrape_google_maps(niche, city, max_results=max_results, headless=headless)
    print(f"Scraped {len(businesses)} businesses from Google Maps.")

    new_rows = []
    for b in businesses:
        key = _dedup_key(b["name"], b["phone"])
        if not key[0]:
            continue
        if key in existing:
            print(f"  skip (dup): {b['name']}")
            continue
        existing.add(key)

        email, source_note = "", ""
        if b["website"]:
            domain = domain_from_url(b["website"])
            email, source_note = hunter_domain_search(domain, hunter_key)
            if not email:
                email, source_note = scrape_website_emails(b["website"])

        qualified, notes = qualify(b["review_count"], source_note)
        row = {
            "Business Name": b["name"],
            "Phone": b["phone"],
            "Email": email,
            "Website": b["website"],
            "City": city,
            "Review Count": b["review_count"],
            "Qualified": qualified,
            "Notes": notes,
        }
        new_rows.append(row)
        print(f"  + {b['name']} | {b['phone']} | {email or '-'} | reviews={b['review_count']} | {qualified}")

    if new_rows:
        write_rows(output, new_rows)
    print(f"Appended {len(new_rows)} new rows to {output}.")


def main():
    ap = argparse.ArgumentParser(description="Google Maps lead scraper")
    ap.add_argument("niche", nargs="?", help='Business niche, e.g. "dental office"')
    ap.add_argument("city", nargs="?", help='City, e.g. "Miami FL"')
    ap.add_argument("--output", default="leads_master.xlsx", help="Master Excel file (appended)")
    ap.add_argument("--max", type=int, default=40, help="Max results to extract")
    ap.add_argument("--hunter-key", default=os.environ.get("HUNTER_API_KEY", ""),
                    help="Hunter.io API key (or set HUNTER_API_KEY env var)")
    ap.add_argument("--no-headless", action="store_true", help="Show browser window")
    args = ap.parse_args()

    niche = args.niche or input("Business niche (e.g. dental office): ").strip()
    city = args.city or input("City (e.g. Miami FL): ").strip()
    if not niche or not city:
        print("Niche and city are required.", file=sys.stderr)
        sys.exit(1)

    if not args.hunter_key:
        print("Warning: no Hunter.io key provided. Falling back to website scraping only.",
              file=sys.stderr)

    run(
        niche=niche,
        city=city,
        output=Path(args.output).expanduser().resolve(),
        hunter_key=args.hunter_key,
        max_results=args.max,
        headless=not args.no_headless,
    )


if __name__ == "__main__":
    main()
